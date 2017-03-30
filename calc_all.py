from __future__ import division

import os

from sklearn.decomposition import PCA

from PyQt4.QtGui import *
import matplotlib.pyplot as plt

import statistics
reload(statistics)
from statistics import *

import openpyxl

from oscore import *

from collections import OrderedDict

import nex

DEFAULT_FILENAME = 'result.xlsx'
PARAMS = dict()   

def get_workbook(fname, indices, data_keys):
    if os.path.isfile(fname):
        wb = openpyxl.load_workbook(fname)
    else:
        wb = openpyxl.Workbook()
    
    row = list(indices)
    row.extend(data_keys)
    print row
    if 'all_results' not in wb.sheetnames:
        ws = wb.create_sheet(0)
        ws.title = 'all_results'
        ws.append(row)
    if 'oscore_results' not in wb.sheetnames:
        ws = wb.create_sheet(1)
        ws.title = 'oscore_results'
        ws.append(row)
    
    wb.save(fname)
    return wb   
    

def write_to_excel(fname, sheet, df, indices):
    data_keys = sorted(list(set(df.keys()) - set(indices)))
    wb = get_workbook(fname, indices, data_keys)
    
    row = [df[ind] for ind in indices]
    row.extend([df[ind] for ind in data_keys])
    
    wb[sheet].append(row)  
    wb.save(fname) 
    
    
def rem_key(data, keys):
    return {k:data[k] for k in data.keys() if k not in keys}


class MyTable(QTableWidget):
    def __init__(self, data, index_header):
        QTableWidget.__init__(self, len(data.values()[0]), len(data))
        self.index_header = index_header
        self.data = data
        self.setmydata()
        self.resizeColumnsToContents()
        self.resizeRowsToContents()
        
 
    def setmydata(self):
        horHeaders = []
        
        self.init_index(horHeaders)
        
        for n, key in enumerate(sorted(rem_key(self.data, self.index_header).keys())):
            horHeaders.append(key)
            for m, item in enumerate(self.data[key]):
                newitem = QTableWidgetItem(str(item))
                self.setItem(m, n+1, newitem)
                
        self.setHorizontalHeaderLabels(horHeaders)
  
    
    def init_index(self, headers):
        headers.append(self.index_header)
        for m, item in enumerate(self.data[self.index_header]):
            newitem = QTableWidgetItem(str(item))
            self.setItem(m, 0, newitem)
        

def draw_table(data):   
    a = QApplication(sys.argv)
 
    table = MyTable(data, 'data_name')
    table.show()
    return a.exec_()
        

def sec_to_timestamps(spikes, freq):
    return np.array((spikes-np.floor(spikes[0]))*PARAMS['frequency'], dtype=np.int32)
	

def get_params():
    global PARAMS
    
    Neuron_Number = 1
    bbh = DEFAULT_BBH
    bbl = DEFAULT_BBL
    brep = DEFAULT_REPEAT
    msize = DEFAULT_WIN_SIZE
    file_path = os.environ['HOMEPATH'] + '\\Desktop\\'
    freq = DEFAULT_FREQ
    fOscillationFreq = 35
    Fmin = 30
    Fmax = 50
    interval = 1
    
    doc = nex.GetActiveDocument()
    __wrapper = []
    res = nex.Dialog(doc, 
                    Neuron_Number, "Select Neuron", "neuron", 
                    interval, "Select interval(s)", "interval",
                    bbh, "Upper bound", "number",
                    bbl, "Lower bound", "number",
                    brep, "Repetition count", "number",
                    msize, "Window size", "number",
					freq, "Ms per bin", "number",
					Fmin, "Min filter frequency", "number",
					Fmax, "Max filter frequency", "number",
                    file_path, "Result file path", "string", 
                    __wrapper )       
    
    Neuron_Number = __wrapper[0]
    Neuron_Var = nex.GetVar(doc, Neuron_Number, "neuron")
    
    interval = __wrapper[1]
    interval_var = nex.GetVar(doc, interval, "interval")
    
    bbh = float(__wrapper[2])
    bbl = float(__wrapper[3])
    brep = int(__wrapper[4])
    msize = int(__wrapper[5])
    freq = float(__wrapper[6])
    fOscillationFreq = 35
    Fmin = int(__wrapper[7])
    Fmax = int(__wrapper[8])
    file_path = str(__wrapper[9])
    
    PARAMS['data'] = np.array(Neuron_Var.Timestamps())
    PARAMS['bbh'] = bbh
    PARAMS['bbl'] = bbl
    PARAMS['brep'] = brep
    PARAMS['msize'] = msize
    PARAMS['filters'] = [p for p in zip(interval_var.Intervals()[0], interval_var.Intervals()[1])]
    PARAMS['data_name'] = nex.GetName(Neuron_Var)
    PARAMS['file_path'] = file_path + '\\' + DEFAULT_FILENAME
    PARAMS['frequency'] = int(1/freq)
    PARAMS['Fmin'] = Fmin
    PARAMS['Fmax'] = Fmax
    PARAMS['doc_name'] = nex.GetDocTitle(doc)


def main():    
    get_params()
    
    data_raw = PARAMS['data']
    data_filtered = np.array([])
    time_int = np.array([])
    interval_len = 0.
    for l, r in PARAMS['filters']:
        interval_len += r - l
         
        spike_data = filter_spikes(data_raw, l, r)
        data_filtered = np.concatenate([data_filtered, spike_data])
        time_int = np.concatenate([time_int, calc_intervals(spike_data)])
            
    if len(time_int) == 0:
        raise 'Empty filter result!'
    
    bi = calc_burst_index(time_int, bbh=PARAMS['bbh'], bbl=PARAMS['bbl'], brep=PARAMS['brep'])
    cv = calc_cv(time_int)
    nu = calc_nu(time_int)
    freq_v = calc_freq_var(time_int)
    mod_burst = calc_modalirity_burst(time_int)
    pause_ind = calc_pause_index(time_int)
    pause_rat = calc_pause_ratio(time_int)
    burst_beh = calc_burst_behavior(time_int)
    skew = calc_skewness(time_int)
    kurt = calc_kurtosis(time_int)
    burst_mean = calc_burst_by_mean(time_int)

    Trial = sec_to_timestamps(data_filtered, PARAMS['frequency']).tolist()
    iTrialLength = Trial[-1]
    oscore = oscore_spikes(np.array([Trial]), iTrialLength, PARAMS['Fmin'], PARAMS['Fmax'], PARAMS['frequency'])
    
    df = dict()
    df['data_name'] = PARAMS['data_name']
    df['burst index'] = bi
    df['cv'] = cv
    df['nu'] = nu
    df['frequency variance'] = freq_v
    df['modalirity burst'] = mod_burst
    df['pause index'] = pause_ind
    df['pause ratio'] = pause_rat
    df['skewness'] = skew
    df['kurtoisis'] = kurt
    df['burst_mean'] = burst_mean
    df['oscore'] = oscore
    df['type'] = get_type(df['burst_mean'], df['cv'])
    df['doc_name'] = PARAMS['doc_name']
    df['isi_mean'] = np.mean(time_int)
    df['isi_median'] = np.median(time_int)
    df['isi_std'] = np.std(time_int)
    df['spike_count'] = len(data_filtered)
    df['filter_length'] = interval_len
    df['bi_2'] = calc_bi_two(data_filtered)
    df['lv'] = calc_local_variance(time_int)
    
    write_to_excel(PARAMS['file_path'], 'all_results', df, ['doc_name', 'data_name'])
    for key in df.keys():
        df[key] = [df[key]]
    
    draw_table(df)
    print 'done'
            
            
if __name__ == '__main__':
    main()