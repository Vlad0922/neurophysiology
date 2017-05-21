# -*- coding: utf-8 -*-

import os
import sys

from PyQt4.QtGui import *
import matplotlib.pyplot as plt

import openpyxl

def get_workbook(fname, indices, data_keys):
    if os.path.isfile(fname):
        wb = openpyxl.load_workbook(fname)
    else:
        wb = openpyxl.Workbook()
    
    row = list(indices)
    row.extend(data_keys)
    if 'all_results' not in wb.sheetnames:
        ws = wb.create_sheet(0)
        ws.title = 'all_results'
        ws.append(row)
    if 'oscore_results' not in wb.sheetnames:
        ws = wb.create_sheet(1)
        ws.title = 'oscore_results'
        ws.append(row)
    
    wb.save(fname)
    return wb   
    

def write_to_excel(fname, sheet, df, indices):
    data_keys = sorted(list(set(df.keys()) - set(indices)))
    wb = get_workbook(fname, indices, data_keys)
    
    row = [df[ind] for ind in indices]
    row.extend([df[ind] for ind in data_keys])
    
    row = map(str, row)
    
    wb[sheet].append(row)  
    wb.save(fname) 
    
    
def rem_key(data, keys):
    return {k:data[k] for k in data.keys() if k not in keys}


class MyTable(QTableWidget):
    def __init__(self, data, index_header):
        QTableWidget.__init__(self, len(data.values()[0]), len(data))
        self.index_header = index_header
        self.data = data
        self.setmydata()
        self.resizeColumnsToContents()
        self.resizeRowsToContents()
        
 
    def setmydata(self):
        horHeaders = []
        
        self.init_index(horHeaders)
        
        for n, key in enumerate(sorted(rem_key(self.data, self.index_header).keys())):
            horHeaders.append(key)
            for m, item in enumerate(self.data[key]):
                newitem = QTableWidgetItem(str(item))
                self.setItem(m, n+1, newitem)
                
        self.setHorizontalHeaderLabels(horHeaders)
  
    
    def init_index(self, headers):
        headers.append(self.index_header)
        for m, item in enumerate(self.data[self.index_header]):
            newitem = QTableWidgetItem(str(item))
            self.setItem(m, 0, newitem)
        

def draw_table(data):   
    a = QApplication(sys.argv)
 
    table = MyTable(data, 'data_name')
    table.show()
    return a.exec_()